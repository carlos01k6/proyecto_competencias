from flask import Blueprint, request, jsonify, send_file
from supabase import Client
from ..supabase_client import get_supabase
import uuid
import jwt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime

excel_bp = Blueprint('excel', __name__, url_prefix='/api/excel')
supabase: Client = get_supabase()

def get_user_id_from_token():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    if not token:
        return None
    try:
        payload = jwt.decode(token, options={"verify_signature": False})
        return payload.get('sub')
    except:
        return None

# EXPORTAR REPORTES A EXCEL
@excel_bp.route('/reporte-evaluacion', methods=['POST'])
def exportar_reporte_evaluacion():
    try:
        usuario_id = get_user_id_from_token()
        if not usuario_id:
            return jsonify({'error': 'No autorizado'}), 401

        data = request.get_json()
        resultado_id = data.get('learning_outcome_id')
        estudiante_id = data.get('student_id')
        nombre_estudiante = data.get('nombre_estudiante', 'Estudiante')
        nombre_resultado = data.get('nombre_resultado', 'Resultado')

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Evaluación'

        # Estilos
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws['A1'] = 'REPORTE DE EVALUACIÓN'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')

        # Información del estudiante
        ws['A3'] = 'Estudiante:'
        ws['B3'] = nombre_estudiante
        ws['A4'] = 'Resultado:'
        ws['B4'] = nombre_resultado
        ws['A5'] = 'Fecha Generación:'
        ws['B5'] = datetime.now().strftime('%d/%m/%Y')

        # Tabla de criterios
        ws['A7'] = 'Criterio'
        ws['B7'] = 'Ponderación'
        ws['C7'] = 'Calificación'
        ws['D7'] = 'Observación'

        for col in ['A7', 'B7', 'C7', 'D7']:
            ws[col].fill = header_fill
            ws[col].font = header_font
            ws[col].border = border

        # Obtener criterios
        criterios_response = supabase.table('criteria').select('*').eq('learning_outcome_id', resultado_id).execute()
        criterios = criterios_response.data

        row = 8
        total_calificado = 0
        suma_ponderada = 0

        for criterio in criterios:
            # Obtener evaluación
            evaluaciones_response = supabase.table('evaluations').select('*').eq('criteria_id', criterio['id']).eq('student_id', estudiante_id).execute()
            
            ws[f'A{row}'] = criterio['name']
            ws[f'B{row}'] = f"{criterio['weighting']}%"
            
            if evaluaciones_response.data:
                evaluacion = evaluaciones_response.data[-1]
                ws[f'C{row}'] = evaluacion['grade']
                ws[f'D{row}'] = evaluacion.get('observation', '')
                total_calificado += 1
                suma_ponderada += evaluacion['grade'] * (criterio['weighting'] / 100)
            else:
                ws[f'C{row}'] = 'Sin calificar'
                ws[f'D{row}'] = ''

            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = border

            row += 1

        # Promedio
        if total_calificado > 0:
            promedio = suma_ponderada / (sum(c['weighting'] for c in criterios) / 100)
        else:
            promedio = 0

        ws[f'A{row + 1}'] = 'CALIFICACIÓN FINAL'
        ws[f'C{row + 1}'] = round(promedio, 2)
        ws[f'A{row + 1}'].font = Font(bold=True)
        ws[f'C{row + 1}'].font = Font(bold=True, size=12)

        # Ajustar columnas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30

        # Guardar en memoria
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'Evaluacion_{nombre_estudiante}_{datetime.now().strftime("%d_%m_%Y")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"ERROR EXPORTAR EXCEL: {str(e)}")
        return jsonify({'error': str(e)}), 500




