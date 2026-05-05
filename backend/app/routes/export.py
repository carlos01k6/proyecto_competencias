from collections import Counter
from datetime import datetime
from io import BytesIO
import json
import traceback

from flask import Blueprint, jsonify, request, send_file
import jwt
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.shapes import Drawing
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from supabase import Client

from ..supabase_client import get_supabase

exportacion_bp = Blueprint("exportacion", __name__, url_prefix="/api/exportar")
supabase: Client = get_supabase()


def get_user_id_from_token():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return None
    try:
        payload = jwt.decode(token, options={"verify_signature": False})
        return payload.get("sub")
    except Exception:
        return None


def obtener_nivel(calificacion):
    calificacion = float(calificacion or 0)
    if calificacion >= 91:
        return "Avanzado"
    if calificacion >= 71:
        return "Satisfactorio"
    if calificacion >= 41:
        return "Basico"
    return "Incipiente"


def encabezado_pie(canvas, doc, titulo):
    canvas.saveState()
    width, height = letter
    canvas.setFont("Helvetica-Bold", 10)
    canvas.drawString(inch, height - 0.5 * inch, "Sistema de Evaluacion por Competencias")
    canvas.setFont("Helvetica", 8)
    canvas.drawString(inch, height - 0.68 * inch, "Reporte institucional")
    canvas.drawRightString(width - inch, height - 0.5 * inch, datetime.now().strftime("%d/%m/%Y"))
    canvas.line(inch, height - 0.78 * inch, width - inch, height - 0.78 * inch)
    canvas.setFont("Helvetica", 8)
    canvas.drawString(inch, 0.5 * inch, titulo)
    canvas.drawRightString(width - inch, 0.5 * inch, f"Pagina {doc.page}")
    canvas.restoreState()


def cargar_periodo(periodo_id):
    if not periodo_id:
        return "Todos los periodos"
    response = supabase.table("configuration").select("key, value").eq("id", periodo_id).execute()
    if response.data:
        config = response.data[0]
        value = config.get("value") or ""
        key = config.get("key") or periodo_id

        try:
            data = json.loads(value)
            nombre = data.get("name") or data.get("description") or key.replace("period_", "")
            fecha_inicio = data.get("start_date")
            fecha_fin = data.get("end_date")

            if fecha_inicio and fecha_fin:
                return f"{nombre} ({fecha_inicio} - {fecha_fin})"
            return nombre
        except Exception:
            return value or key
    return periodo_id


def obtener_evaluaciones_por_criterios(criteria_ids, student_id=None):
    if not criteria_ids:
        return []
    query = supabase.table("evaluations").select("criteria_id, student_id, grade, grading_date, created_at").in_("criteria_id", criteria_ids)
    if student_id:
        query = query.eq("student_id", student_id)
    response = query.execute()
    evaluaciones = [
        evaluation
        for evaluation in response.data or []
        if evaluation.get("grade") is not None
    ]
    ultimas = {}
    for evaluation in evaluaciones:
        key = (evaluation.get("student_id"), evaluation.get("criteria_id"))
        fecha = evaluation.get("grading_date") or evaluation.get("created_at") or ""
        actual = ultimas.get(key)
        if not actual or fecha >= (actual.get("grading_date") or actual.get("created_at") or ""):
            ultimas[key] = evaluation
    return list(ultimas.values())


def obtener_datos_reporte(tipo, periodo_id=None):
    tipo = tipo if tipo in {"competencia", "grado", "asignatura"} else "competencia"
    periodo_nombre = cargar_periodo(periodo_id)

    competencias_response = supabase.table("competencies").select("*").execute()
    competencias = competencias_response.data or []

    filas_competencia = []
    todas_las_notas = []

    for competencia in competencias:
        resultados_response = (
            supabase.table("learning_outcomes")
            .select("id")
            .eq("competency_id", competencia.get("id"))
            .execute()
        )
        resultado_ids = [resultado["id"] for resultado in resultados_response.data or []]

        criterios = []
        if resultado_ids:
            criterios_response = (
                supabase.table("criteria")
                .select("id")
                .in_("learning_outcome_id", resultado_ids)
                .execute()
            )
            criterios = criterios_response.data or []

        criterio_ids = [criterio["id"] for criterio in criterios]
        evaluaciones = obtener_evaluaciones_por_criterios(criterio_ids)
        notas = [float(evaluacion["grade"]) for evaluacion in evaluaciones]
        todas_las_notas.extend(notas)

        if notas:
            promedio = sum(notas) / len(notas)
            filas_competencia.append(
                {
                    "grupo": competencia.get("name") or competencia.get("nombre") or competencia.get("id"),
                    "descripcion": competencia.get("description") or "",
                    "total_evaluaciones": len(notas),
                    "promedio": round(promedio, 2),
                    "nivel": obtener_nivel(promedio),
                }
            )

    if tipo == "competencia":
        filas = filas_competencia
    else:
        agrupadas = {}
        campo = "subject" if tipo == "asignatura" else "grade"
        etiqueta_default = "Sin asignatura" if tipo == "asignatura" else "Sin grado"
        for competencia, fila in zip(competencias, filas_competencia):
            clave = competencia.get(campo) or competencia.get("grado") or etiqueta_default
            if clave not in agrupadas:
                agrupadas[clave] = {"notas": [], "total": 0}
            agrupadas[clave]["total"] += fila["total_evaluaciones"]
            if fila["total_evaluaciones"]:
                agrupadas[clave]["notas"].append(fila["promedio"])

        filas = []
        for clave, datos in agrupadas.items():
            promedio = sum(datos["notas"]) / len(datos["notas"]) if datos["notas"] else 0
            filas.append(
                {
                    "grupo": clave,
                    "descripcion": "",
                    "total_evaluaciones": datos["total"],
                    "promedio": round(promedio, 2),
                    "nivel": obtener_nivel(promedio),
                }
            )

    distribucion = Counter(obtener_nivel(nota) for nota in todas_las_notas)
    resumen = {
        "total_registros": len(filas),
        "total_evaluaciones": len(todas_las_notas),
        "promedio_general": round(sum(todas_las_notas) / len(todas_las_notas), 2) if todas_las_notas else 0,
        "periodo": periodo_nombre,
        "tipo": tipo,
    }

    return {
        "titulo": f"Reporte institucional por {tipo}",
        "tipo": tipo,
        "periodo_id": periodo_id,
        "periodo": periodo_nombre,
        "filas": filas,
        "distribucion": dict(distribucion),
        "resumen": resumen,
    }


def obtener_boletin_estudiante(student_id):
    estudiante_response = supabase.table("users").select("id, name, email").eq("id", student_id).execute()
    estudiante = estudiante_response.data[0] if estudiante_response.data else {"id": student_id, "name": "Estudiante", "email": ""}

    competencias_response = supabase.table("competencies").select("*").execute()
    competencias = competencias_response.data or []
    filas = []
    notas_generales = []

    for competencia in competencias:
        resultados_response = (
            supabase.table("learning_outcomes")
            .select("id")
            .eq("competency_id", competencia.get("id"))
            .execute()
        )
        resultado_ids = [resultado["id"] for resultado in resultados_response.data or []]
        if not resultado_ids:
            continue

        criterios_response = (
            supabase.table("criteria")
            .select("id")
            .in_("learning_outcome_id", resultado_ids)
            .execute()
        )
        criterio_ids = [criterio["id"] for criterio in criterios_response.data or []]
        evaluaciones = obtener_evaluaciones_por_criterios(criterio_ids, student_id)
        notas = [float(evaluacion["grade"]) for evaluacion in evaluaciones]
        if not notas:
            continue

        promedio = sum(notas) / len(notas)
        notas_generales.extend(notas)
        filas.append(
            {
                "competencia": competencia.get("name") or competencia.get("nombre") or competencia.get("id"),
                "total_evaluaciones": len(notas),
                "promedio": round(promedio, 2),
                "nivel": obtener_nivel(promedio),
            }
        )

    promedio_general = sum(notas_generales) / len(notas_generales) if notas_generales else 0
    return {
        "estudiante": estudiante,
        "filas": filas,
        "resumen": {
            "promedio_general": round(promedio_general, 2),
            "nivel_general": obtener_nivel(promedio_general),
            "total_competencias": len(filas),
            "total_evaluaciones": len(notas_generales),
        },
    }


def agregar_tabla_pdf(story, filas):
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle(
        "TableCell",
        parent=styles["BodyText"],
        fontSize=8,
        leading=10,
        wordWrap="CJK",
    )
    header_style = ParagraphStyle(
        "TableHeader",
        parent=cell_style,
        textColor=colors.white,
        fontName="Helvetica-Bold",
    )

    data = [[
        Paragraph("Nombre", header_style),
        Paragraph("Descripcion", header_style),
        Paragraph("Evaluaciones", header_style),
        Paragraph("Promedio", header_style),
        Paragraph("Nivel", header_style),
    ]]
    for fila in filas:
        data.append([
            Paragraph(str(fila.get("grupo") or fila.get("competencia") or ""), cell_style),
            Paragraph(str(fila.get("descripcion", ""))[:120], cell_style),
            fila.get("total_evaluaciones", 0),
            fila.get("promedio", 0),
            Paragraph(str(fila.get("nivel", "")), cell_style),
        ])

    table = Table(
        data,
        colWidths=[1.45 * inch, 3.05 * inch, 0.8 * inch, 0.7 * inch, 0.9 * inch],
        repeatRows=1,
        hAlign="LEFT",
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d1d5db")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(table)


def agregar_grafico_pdf(story, distribucion):
    labels = ["Avanzado", "Satisfactorio", "Basico", "Incipiente"]
    valores = [distribucion.get(label, 0) for label in labels]
    drawing = Drawing(420, 220)
    chart = VerticalBarChart()
    chart.x = 50
    chart.y = 40
    chart.height = 150
    chart.width = 330
    chart.data = [valores]
    chart.categoryAxis.categoryNames = labels
    chart.valueAxis.valueMin = 0
    chart.bars[0].fillColor = colors.HexColor("#2563eb")
    drawing.add(chart)
    story.append(drawing)


def crear_pdf_reporte(datos):
    buffer = BytesIO()
    titulo = datos["titulo"]
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.0 * inch, bottomMargin=0.8 * inch)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Institucion Educativa", styles["Title"]),
        Paragraph(titulo, styles["Heading1"]),
        Paragraph(f"Periodo: {datos['periodo']}", styles["Normal"]),
        Spacer(1, 0.2 * inch),
    ]
    agregar_tabla_pdf(story, datos["filas"])
    story.append(Spacer(1, 0.3 * inch))
    story.append(Paragraph("Distribucion de niveles", styles["Heading2"]))
    agregar_grafico_pdf(story, datos["distribucion"])
    doc.build(story, onFirstPage=lambda c, d: encabezado_pie(c, d, titulo), onLaterPages=lambda c, d: encabezado_pie(c, d, titulo))
    buffer.seek(0)
    return buffer


def crear_excel_reporte(datos):
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    headers = ["Nombre", "Descripcion", "Evaluaciones", "Promedio", "Nivel"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")

    for fila in datos["filas"]:
        ws.append([
            fila.get("grupo") or fila.get("competencia"),
            fila.get("descripcion", ""),
            fila.get("total_evaluaciones", 0),
            fila.get("promedio", 0),
            fila.get("nivel", ""),
        ])

    ws_chart = wb.create_sheet("Graficos")
    ws_chart.append(["Nivel", "Cantidad"])
    for nivel in ["Avanzado", "Satisfactorio", "Basico", "Incipiente"]:
        ws_chart.append([nivel, datos["distribucion"].get(nivel, 0)])
    chart = BarChart()
    chart.title = "Distribucion de niveles"
    chart.y_axis.title = "Cantidad"
    chart.x_axis.title = "Nivel"
    data = Reference(ws_chart, min_col=2, min_row=1, max_row=5)
    cats = Reference(ws_chart, min_col=1, min_row=2, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_chart.add_chart(chart, "D2")

    ws_summary = wb.create_sheet("Resumen")
    for key, value in datos["resumen"].items():
        ws_summary.append([key, value])
    for row in wb.worksheets:
        for column in row.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            row.column_dimensions[column[0].column_letter].width = min(max_length + 2, 45)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def crear_pdf_boletin(datos):
    buffer = BytesIO()
    estudiante = datos["estudiante"]
    titulo = f"Boletin individual - {estudiante.get('name') or estudiante.get('email') or estudiante.get('id')}"
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.0 * inch, bottomMargin=0.8 * inch)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Institucion Educativa", styles["Title"]),
        Paragraph(titulo, styles["Heading1"]),
        Paragraph(f"Promedio general: {datos['resumen']['promedio_general']} - {datos['resumen']['nivel_general']}", styles["Normal"]),
        Spacer(1, 0.2 * inch),
    ]
    agregar_tabla_pdf(story, datos["filas"])
    doc.build(story, onFirstPage=lambda c, d: encabezado_pie(c, d, titulo), onLaterPages=lambda c, d: encabezado_pie(c, d, titulo))
    buffer.seek(0)
    return buffer


def crear_excel_boletin(datos):
    wb = Workbook()
    ws = wb.active
    ws.title = "Boletin"
    estudiante = datos["estudiante"]
    ws.append(["Estudiante", estudiante.get("name") or estudiante.get("email") or estudiante.get("id")])
    ws.append(["Promedio general", datos["resumen"]["promedio_general"]])
    ws.append(["Nivel general", datos["resumen"]["nivel_general"]])
    ws.append([])
    ws.append(["Competencia", "Evaluaciones", "Promedio", "Nivel"])
    for fila in datos["filas"]:
        ws.append([fila["competencia"], fila["total_evaluaciones"], fila["promedio"], fila["nivel"]])

    ws_chart = wb.create_sheet("Graficos")
    dist = Counter(fila["nivel"] for fila in datos["filas"])
    ws_chart.append(["Nivel", "Cantidad"])
    for nivel in ["Avanzado", "Satisfactorio", "Basico", "Incipiente"]:
        ws_chart.append([nivel, dist.get(nivel, 0)])
    chart = BarChart()
    chart.title = "Niveles por competencia"
    data = Reference(ws_chart, min_col=2, min_row=1, max_row=5)
    cats = Reference(ws_chart, min_col=1, min_row=2, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_chart.add_chart(chart, "D2")

    ws_summary = wb.create_sheet("Resumen")
    for key, value in datos["resumen"].items():
        ws_summary.append([key, value])
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 45)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@exportacion_bp.route("/reportes/preview", methods=["GET"])
def preview_reporte():
    try:
        if not get_user_id_from_token():
            return jsonify({"error": "No autorizado"}), 401
        datos = obtener_datos_reporte(request.args.get("tipo", "competencia"), request.args.get("periodo_id"))
        return jsonify(datos), 200
    except Exception as e:
        print(f"ERROR PREVIEW EXPORTACION: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@exportacion_bp.route("/reportes/pdf", methods=["GET"])
def exportar_reportes_pdf():
    try:
        if not get_user_id_from_token():
            return jsonify({"error": "No autorizado"}), 401
        datos = obtener_datos_reporte(request.args.get("tipo", "competencia"), request.args.get("periodo_id"))
        buffer = crear_pdf_reporte(datos)
        return send_file(buffer, as_attachment=True, download_name="reporte_institucional.pdf", mimetype="application/pdf")
    except Exception as e:
        print(f"ERROR EXPORTAR PDF: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@exportacion_bp.route("/reportes/excel", methods=["GET"])
def exportar_reportes_excel():
    try:
        if not get_user_id_from_token():
            return jsonify({"error": "No autorizado"}), 401
        datos = obtener_datos_reporte(request.args.get("tipo", "competencia"), request.args.get("periodo_id"))
        buffer = crear_excel_reporte(datos)
        return send_file(
            buffer,
            as_attachment=True,
            download_name="reporte_institucional.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        print(f"ERROR EXPORTAR EXCEL: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@exportacion_bp.route("/boletines/pdf/<student_id>", methods=["GET"])
def exportar_boletin_pdf(student_id):
    try:
        if not get_user_id_from_token():
            return jsonify({"error": "No autorizado"}), 401
        buffer = crear_pdf_boletin(obtener_boletin_estudiante(student_id))
        return send_file(buffer, as_attachment=True, download_name=f"boletin_{student_id}.pdf", mimetype="application/pdf")
    except Exception as e:
        print(f"ERROR EXPORTAR BOLETIN PDF: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@exportacion_bp.route("/boletines/excel/<student_id>", methods=["GET"])
def exportar_boletin_excel(student_id):
    try:
        if not get_user_id_from_token():
            return jsonify({"error": "No autorizado"}), 401
        buffer = crear_excel_boletin(obtener_boletin_estudiante(student_id))
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"boletin_{student_id}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        print(f"ERROR EXPORTAR BOLETIN EXCEL: {str(e)}")
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
